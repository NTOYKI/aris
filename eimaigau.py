# -*- coding: utf-8 -*-
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

wb = Workbook()

FONT_NAME = "Arial"
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="FFFFFF")
HEAD_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
SUBHEAD_FONT = Font(name=FONT_NAME, bold=True, color="000000")
BLUE = Font(name=FONT_NAME, color="0000FF")
BLUE_B = Font(name=FONT_NAME, bold=True, color="0000FF")
BLACK = Font(name=FONT_NAME, color="000000")
GREEN = Font(name=FONT_NAME, color="008000")
TITLE_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
HEAD_FILL = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
LIGHT_FILL = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")
CLOSE_FILL = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")
REST_FILL = PatternFill("solid", start_color="E7E6E6", end_color="E7E6E6")
SHORTAGE_FILL = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
SHORTAGE_FONT = Font(name=FONT_NAME, color="9C0006")
OK_FILL = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BEACH = "Μπουσουλάς"
PEOPLE = ["Γιάννης", "Γιώτα", "Ραφαήλ", "Άρης", "Οδυσσέας", "Κώστας"]
ROTATORS = ["Γιώτα", "Ραφαήλ", "Άρης", "Κώστας"]
DAY_NAMES = ["Δευτέρα", "Τρίτη", "Τετάρτη", "Πέμπτη", "Παρασκευή", "Σάββατο", "Κυριακή"]
DAY_COLS = list(range(3, 10))  # C..I
FIRST_DAY_COL = 3
FIRST_SHIFT_ROW = 12
NATURAL_SHIFT_ROW = FIRST_SHIFT_ROW
EFFECTIVE_SHIFT_ROW = FIRST_SHIFT_ROW + 8
OPENING_COUNT_ROW = FIRST_SHIFT_ROW + 6
POSITION_START_ROW = EFFECTIVE_SHIFT_ROW + 8
HOURS_START_ROW = POSITION_START_ROW + 8
CHECK_ROW = HOURS_START_ROW + 10

DEFAULT_REST = {
    "Γιάννης": "Κυριακή",
    "Γιώτα": "Δευτέρα",
    "Ραφαήλ": "Τρίτη",
    "Άρης": "Τετάρτη",
    "Οδυσσέας": "Πέμπτη",
    "Κώστας": "Παρασκευή",
}

PARAM_PEOPLE_START = 7
PARAM_INFINITY_COVERAGE = 14
PARAM_BACKUP_OPENER = 15
PARAM_OPEN_HOURS = 17
PARAM_CLOSE_HOURS = 18
PARAM_GIANNIS_HOURS = 19

SECTION_HEADINGS = {
    "ΤΙ ΦΤΙΑΞΑΜΕ",
    "ΚΑΝΟΝΕΣ ΠΟΥ ΕΧΟΥΝ ΚΩΔΙΚΟΠΟΙΗΘΕΙ",
    "ΧΡΩΜΑΤΙΚΟΣ ΚΩΔΙΚΑΣ ΣΤΟΝ ΒΟΗΘΗΤΙΚΟ ΠΙΝΑΚΑ ΩΡΩΝ",
    "ΣΗΜΑΝΤΙΚΗ ΕΠΙΣΗΜΑΝΣΗ ΓΙΑ ΤΙΣ ΜΕΡΕΣ ΧΩΡΙΣ ΦΥΣΙΚΟ «ΑΝΟΙΓΜΑ»",
    "ΤΙ ΝΑ ΣΥΜΠΛΗΡΩΣΕΙΣ (μπλε κελιά στο φύλλο «Παράμετροι»)",
}


def style_range(ws, min_row, max_row, min_col, max_col, font=None, fill=None, alignment=None, border=None):
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if alignment:
                cell.alignment = alignment
            if border:
                cell.border = border


def next_day_col(col):
    return col + 1 if col < 9 else FIRST_DAY_COL


def day_col_letter(col):
    return get_column_letter(col)


def param_rest_ref(person_row_index):
    return f"Παράμετροι!$B${PARAM_PEOPLE_START + person_row_index}"


def person_name_cell(row):
    return f"$A{row}"


def natural_shift_formula(row, person_index, col):
    rest_ref = param_rest_ref(person_index)
    cur = f"{day_col_letter(col)}$3"
    nxt = f"{day_col_letter(next_day_col(col))}$3"
    return f'=IF({cur}={rest_ref},"Ρεπό",IF({nxt}={rest_ref},"Άνοιγμα","Κλείσιμο"))'


def effective_shift_formula(row, person_index, col):
    rest_ref = param_rest_ref(person_index)
    cur = f"{day_col_letter(col)}$3"
    nxt = f"{day_col_letter(next_day_col(col))}$3"
    natural_row = NATURAL_SHIFT_ROW + person_index
    count_row = OPENING_COUNT_ROW
    col_letter = day_col_letter(col)
    return (
        f'=IF({cur}={rest_ref},"Ρεπό",IF({nxt}={rest_ref},"Άνοιγμα",'
        f'IF(AND({col_letter}${count_row}=0,{person_name_cell(row)}=Παράμετροι!$B${PARAM_BACKUP_OPENER},'
        f'{col_letter}{natural_row}="Κλείσιμο"),"Άνοιγμα",{col_letter}{natural_row})))'
    )


def coverage_shift_formula(col):
    col_letter = day_col_letter(col)
    return (
        f'=IFERROR(INDEX({col_letter}{EFFECTIVE_SHIFT_ROW}:{col_letter}{EFFECTIVE_SHIFT_ROW + 5},'
        f'MATCH(Παράμετροι!$B${PARAM_INFINITY_COVERAGE},$A{EFFECTIVE_SHIFT_ROW}:$A{EFFECTIVE_SHIFT_ROW + 5},0)),"Ρεπό")'
    )


def beach_formula(col):
    col_letter = day_col_letter(col)
    giannis = EFFECTIVE_SHIFT_ROW
    odysseas = EFFECTIVE_SHIFT_ROW + 4
    return (
        f'=IF({col_letter}{giannis}<>"Ρεπό","Γιάννης",'
        f'IF({col_letter}{odysseas}<>"Ρεπό","Οδυσσέας",""))'
    )


def infinity_formula(col):
    col_letter = day_col_letter(col)
    giannis = EFFECTIVE_SHIFT_ROW
    odysseas = EFFECTIVE_SHIFT_ROW + 4
    cov_shift = coverage_shift_formula(col)
    return (
        f'=IF(AND({col_letter}{giannis}<>"Ρεπό",{col_letter}{odysseas}<>"Ρεπό"),"Οδυσσέας",'
        f'IF({cov_shift}<>"Ρεπό",Παράμετροι!$B${PARAM_INFINITY_COVERAGE},""))'
    )


def rotator_pool_formula(person_name, rotator_index, col):
    col_letter = day_col_letter(col)
    person_row = EFFECTIVE_SHIFT_ROW + PEOPLE.index(person_name)
    return (
        f'=IF({col_letter}{person_row}="Ρεπό","",'
        f'IF(MOD(WEEKDAY({col_letter}$4,2)-1+{rotator_index},2)=0,"Main Pool","Splash Pool"))'
    )


def pool_name_formula(col, pool_name, helper_start_row):
    parts = []
    for i in range(4):
        helper_row = helper_start_row + i
        parts.append(f'IF({day_col_letter(col)}{helper_row}="{pool_name}",$A{helper_row},"")')
    return f'=TEXTJOIN(" / ",TRUE,{",".join(parts)})'


def hours_formula(person_name, person_index, col):
    col_letter = day_col_letter(col)
    effective_row = EFFECTIVE_SHIFT_ROW + person_index
    natural_row = NATURAL_SHIFT_ROW + person_index
    if person_name == "Γιάννης":
        return (
            f'=IF({col_letter}{effective_row}="Ρεπό","",'
            f'IF({col_letter}{natural_row}="Άνοιγμα",Παράμετροι!$B${PARAM_OPEN_HOURS},'
            f'Παράμετροι!$B${PARAM_GIANNIS_HOURS}))'
        )
    return (
        f'=IF({col_letter}{effective_row}="Ρεπό","",'
        f'IF({col_letter}{effective_row}="Άνοιγμα",Παράμετροι!$B${PARAM_OPEN_HOURS},'
        f'Παράμετροι!$B${PARAM_CLOSE_HOURS}))'
    )


# ==================================================================
# SHEET 0: ΟΔΗΓΙΕΣ
# ==================================================================
ws0 = wb.active
ws0.title = "Οδηγίες"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 105

ws0["A1"] = "ΠΡΟΓΡΑΜΜΑ ΝΑΥΑΓΟΣΩΣΤΩΝ – Οδηγίες"
ws0["A1"].font = TITLE_FONT
ws0["A1"].fill = TITLE_FILL
ws0.row_dimensions[1].height = 26

lines = [
    "",
    "ΤΙ ΦΤΙΑΞΑΜΕ",
    "Ένα εβδομαδιαίο πρόγραμμα για 6 ναυαγοσώστες: Γιάννης, Γιώτα, Ραφαήλ, Άρης, Οδυσσέας, Κώστας — σε 4 θέσεις: "
    f"{BEACH}, Infinity Pool, Main Pool, Splash Pool. Οι ημερομηνίες ενημερώνονται μόνες τους κάθε εβδομάδα "
    "(ξεκινούν πάντα από την επόμενη/τρέχουσα Δευτέρα).",
    "",
    "ΚΑΝΟΝΕΣ ΠΟΥ ΕΧΟΥΝ ΚΩΔΙΚΟΠΟΙΗΘΕΙ",
    f"• Ο Γιάννης είναι πάντα στον {BEACH} όποτε δουλεύει.",
    "• Ο Οδυσσέας είναι πάντα στο Infinity Pool όποτε δουλεύει.",
    "• Καθένας δουλεύει 6 μέρες/εβδομάδα, δηλ. έχει 1 ρεπό/εβδομάδα (ρυθμίζεται στο φύλλο «Παράμετροι»).",
    "• Την ημέρα ΠΡΙΝ το δικό του ρεπό, ο καθένας κάνει «Άνοιγμα» (πρωινή βάρδια 9:00-17:30). Τις υπόλοιπες "
    "εργάσιμες μέρες κάνει «Κλείσιμο» (10:30-19:00).",
    f"• Εξαίρεση: τις κανονικές του μέρες (όχι πριν το ρεπό) ο Γιάννης δουλεύει 9:30-18:00 στον {BEACH}.",
    f"• Στο ρεπό του Γιάννη, ο Οδυσσέας καλύπτει τον {BEACH} (κι έτσι το Infinity Pool μένει χωρίς τον Οδυσσέα "
    "εκείνη τη μέρα).",
    "• Όποτε ο Οδυσσέας λείπει από το Infinity Pool (ρεπό ή κάλυψη παραλίας), το Infinity Pool καλύπτεται από "
    "τον ναυαγοσώστη που έχεις ορίσει στο πεδίο «Κάλυψη Infinity Pool».",
    "• Οι υπόλοιποι (Γιώτα, Ραφαήλ, Άρης, Κώστας) εναλλάσσονται στο Main Pool και στο Splash Pool.",
    "• Απαιτούνται τουλάχιστον 1 άτομο σε βάρδια «Άνοιγμα» και τουλάχιστον 3 άτομα σε βάρδια «Κλείσιμο», κάθε "
    "μέρα — ελέγχεται αυτόματα στο κάτω μέρος του φύλλου «Εβδομαδιαίο Πρόγραμμα».",
    "",
    "ΧΡΩΜΑΤΙΚΟΣ ΚΩΔΙΚΑΣ ΣΤΟΝ ΒΟΗΘΗΤΙΚΟ ΠΙΝΑΚΑ ΩΡΩΝ",
    "• Γαλάζιο = βάρδια Ανοίγματος. Ροζ/πορτοκαλί = βάρδια Κλεισίματος. Γκρι = Ρεπό.",
    "",
    "ΣΗΜΑΝΤΙΚΗ ΕΠΙΣΗΜΑΝΣΗ ΓΙΑ ΤΙΣ ΜΕΡΕΣ ΧΩΡΙΣ ΦΥΣΙΚΟ «ΑΝΟΙΓΜΑ»",
    "Με 6 άτομα σε 7 μέρες, κάποιες μέρες της εβδομάδας μπορεί να μην έχει κανείς ρεπό την επόμενη μέρα, οπότε "
    "κανείς δεν θα άνοιγε φυσικά. Σε αυτές τις μέρες ανοίγει αυτόματα ο «Εφεδρικός ανοίγματος» που έχεις ορίσει "
    "στις Παραμέτρους (προεπιλογή: Κώστας). Αν δεις ότι ανοίγει πολύ συχνά, σκέψου να αλλάξεις τα ρεπό ή τον "
    "εφεδρικό ώστε να μοιράζεται πιο δίκαια.",
    "",
    "ΤΙ ΝΑ ΣΥΜΠΛΗΡΩΣΕΙΣ (μπλε κελιά στο φύλλο «Παράμετροι»)",
    "1. Ρεπό κάθε ατόμου (dropdown με μέρες εβδομάδας).",
    "2. Ποιος καλύπτει το Infinity Pool στην απουσία του Οδυσσέα.",
    "3. Ποιος είναι ο εφεδρικός ανοίγματος για μέρες χωρίς φυσικό ανοιχτή.",
    "4. Τις ώρες (άνοιγμα/κλείσιμο/ωράριο Γιάννη) αν αλλάξουν ποτέ.",
    "",
    "Αν κάτι δεν σε καλύπτει, μπορείς να γράψεις χειροκίνητα πάνω σε οποιοδήποτε κελί του «Εβδομαδιαίου "
    "Προγράμματος».",
]
r = 2
for line in lines:
    cell = ws0.cell(row=r, column=1, value=line)
    if line in SECTION_HEADINGS:
        cell.font = Font(name=FONT_NAME, bold=True, size=12, color="1F4E78")
    else:
        cell.font = Font(name=FONT_NAME, size=10.5)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws0.row_dimensions[r].height = 30 if len(line) > 95 else 16
    r += 1

# ==================================================================
# SHEET 1: ΠΑΡΑΜΕΤΡΟΙ
# ==================================================================
ws1 = wb.create_sheet("Παράμετροι")
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 34
ws1.column_dimensions["B"].width = 22

ws1["A1"] = "ΠΑΡΑΜΕΤΡΟΙ ΠΡΟΓΡΑΜΜΑΤΟΣ"
ws1["A1"].font = TITLE_FONT
ws1["A1"].fill = TITLE_FILL
ws1.merge_cells("A1:B1")
ws1.row_dimensions[1].height = 26

ws1["A3"] = "Εβδομάδα από (Δευτέρα)"
ws1["A3"].font = SUBHEAD_FONT
ws1["B3"] = "=TODAY()-WEEKDAY(TODAY(),2)+1"
ws1["B3"].number_format = "dd/mm/yyyy"
ws1["B3"].font = BLUE_B

ws1["A6"] = "Ναυαγοσώστης"
ws1["B6"] = "Ρεπό (ημέρα)"
style_range(ws1, 6, 6, 1, 2, font=HEAD_FONT, fill=HEAD_FILL, alignment=Alignment(horizontal="center"), border=BORDER)

for i, name in enumerate(PEOPLE):
    row = PARAM_PEOPLE_START + i
    ws1.cell(row=row, column=1, value=name).font = BLACK
    rest_cell = ws1.cell(row=row, column=2, value=DEFAULT_REST[name])
    rest_cell.font = BLUE
    for col in (1, 2):
        ws1.cell(row=row, column=col).border = BORDER
        ws1.cell(row=row, column=col).alignment = Alignment(horizontal="center")

ws1.cell(row=PARAM_INFINITY_COVERAGE, column=1, value="Κάλυψη Infinity Pool").font = SUBHEAD_FONT
ws1.cell(row=PARAM_INFINITY_COVERAGE, column=2, value="Ραφαήλ").font = BLUE
ws1.cell(row=PARAM_BACKUP_OPENER, column=1, value="Εφεδρικός Ανοίγματος").font = SUBHEAD_FONT
ws1.cell(row=PARAM_BACKUP_OPENER, column=2, value="Κώστας").font = BLUE

ws1.cell(row=PARAM_OPEN_HOURS, column=1, value="Ώρες Άνοιγμα").font = SUBHEAD_FONT
ws1.cell(row=PARAM_OPEN_HOURS, column=2, value="9:00-17:30").font = BLUE
ws1.cell(row=PARAM_CLOSE_HOURS, column=1, value="Ώρες Κλείσιμο").font = SUBHEAD_FONT
ws1.cell(row=PARAM_CLOSE_HOURS, column=2, value="10:30-19:00").font = BLUE
ws1.cell(row=PARAM_GIANNIS_HOURS, column=1, value="Γιάννης (κανονικά)").font = SUBHEAD_FONT
ws1.cell(row=PARAM_GIANNIS_HOURS, column=2, value="9:30-18:00").font = BLUE

day_list = ",".join(DAY_NAMES)
people_list = ",".join(PEOPLE)

dv_days = DataValidation(type="list", formula1=f'"{day_list}"', allow_blank=False)
dv_people = DataValidation(type="list", formula1=f'"{people_list}"', allow_blank=False)
ws1.add_data_validation(dv_days)
ws1.add_data_validation(dv_people)

for row in range(PARAM_PEOPLE_START, PARAM_PEOPLE_START + len(PEOPLE)):
    dv_days.add(ws1.cell(row=row, column=2))
dv_people.add(ws1.cell(row=PARAM_INFINITY_COVERAGE, column=2))
dv_people.add(ws1.cell(row=PARAM_BACKUP_OPENER, column=2))

# ==================================================================
# SHEET 2: ΕΒΔΟΜΑΔΙΑΙΟ ΠΡΟΓΡΑΜΜΑ
# ==================================================================
ws2 = wb.create_sheet("Εβδομαδιαίο Πρόγραμμα")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 14
for col in DAY_COLS:
    ws2.column_dimensions[get_column_letter(col)].width = 16

ws2["A1"] = "ΕΒΔΟΜΑΔΙΑΙΟ ΠΡΟΓΡΑΜΜΑ"
ws2["A1"].font = TITLE_FONT
ws2["A1"].fill = TITLE_FILL
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
ws2.row_dimensions[1].height = 26

ws2["A3"] = "Ημέρα"
ws2["B3"] = "Ημ/νία"
for idx, day_name in enumerate(DAY_NAMES):
    col = FIRST_DAY_COL + idx
    ws2.cell(row=3, column=col, value=day_name)
    date_cell = ws2.cell(row=4, column=col)
    if idx == 0:
        date_cell.value = "=Παράμετροι!$B$3"
    else:
        prev = get_column_letter(col - 1)
        date_cell.value = f"={prev}4+1"
    date_cell.number_format = "dd/mm/yyyy"
style_range(ws2, 3, 4, 1, 9, font=HEAD_FONT, fill=HEAD_FILL, alignment=Alignment(horizontal="center"), border=BORDER)

ws2.cell(row=FIRST_SHIFT_ROW - 1, column=1, value="ΦΥΣΙΚΕΣ ΒΑΡΔΙΕΣ").font = SUBHEAD_FONT
ws2.merge_cells(start_row=FIRST_SHIFT_ROW - 1, start_column=1, end_row=FIRST_SHIFT_ROW - 1, end_column=2)

for i, name in enumerate(PEOPLE):
    row = NATURAL_SHIFT_ROW + i
    ws2.cell(row=row, column=1, value=name).font = BLACK
    ws2.cell(row=row, column=1).border = BORDER
    for col in DAY_COLS:
        cell = ws2.cell(row=row, column=col)
        cell.value = natural_shift_formula(row, i, col)
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

for col in DAY_COLS:
    col_letter = day_col_letter(col)
    ws2.cell(row=OPENING_COUNT_ROW, column=col, value=f'=COUNTIF({col_letter}{NATURAL_SHIFT_ROW}:{col_letter}{NATURAL_SHIFT_ROW + 5},"Άνοιγμα")')
    ws2.cell(row=OPENING_COUNT_ROW, column=col).alignment = Alignment(horizontal="center")

ws2.cell(row=EFFECTIVE_SHIFT_ROW - 1, column=1, value="ΤΕΛΙΚΕΣ ΒΑΡΔΙΕΣ").font = SUBHEAD_FONT
ws2.merge_cells(start_row=EFFECTIVE_SHIFT_ROW - 1, start_column=1, end_row=EFFECTIVE_SHIFT_ROW - 1, end_column=2)

for i, name in enumerate(PEOPLE):
    row = EFFECTIVE_SHIFT_ROW + i
    ws2.cell(row=row, column=1, value=name).font = BLACK
    ws2.cell(row=row, column=1).border = BORDER
    for col in DAY_COLS:
        cell = ws2.cell(row=row, column=col)
        cell.value = effective_shift_formula(row, i, col)
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

positions = [BEACH, "Infinity Pool", "Main Pool", "Splash Pool"]
ws2.cell(row=POSITION_START_ROW - 1, column=1, value="ΘΕΣΕΙΣ ΕΡΓΑΣΙΑΣ").font = SUBHEAD_FONT
ws2.merge_cells(start_row=POSITION_START_ROW - 1, start_column=1, end_row=POSITION_START_ROW - 1, end_column=2)

pool_helper_start = POSITION_START_ROW + len(positions) + 2
for i, rotator in enumerate(ROTATORS):
    row = pool_helper_start + i
    ws2.cell(row=row, column=1, value=rotator).font = BLACK
    for col in DAY_COLS:
        ws2.cell(row=row, column=col, value=rotator_pool_formula(rotator, i, col))

for i, position in enumerate(positions):
    row = POSITION_START_ROW + i
    ws2.cell(row=row, column=1, value=position).font = BLACK
    ws2.cell(row=row, column=1).border = BORDER
    for col in DAY_COLS:
        cell = ws2.cell(row=row, column=col)
        if position == BEACH:
            cell.value = beach_formula(col)
        elif position == "Infinity Pool":
            cell.value = infinity_formula(col)
        else:
            cell.value = pool_name_formula(col, position, pool_helper_start)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BORDER

ws2.cell(row=HOURS_START_ROW - 1, column=1, value="ΒΟΗΘΗΤΙΚΟΣ ΠΙΝΑΚΑΣ ΩΡΩΝ").font = SUBHEAD_FONT
ws2.merge_cells(start_row=HOURS_START_ROW - 1, start_column=1, end_row=HOURS_START_ROW - 1, end_column=2)

for i, name in enumerate(PEOPLE):
    row = HOURS_START_ROW + i
    ws2.cell(row=row, column=1, value=name).font = BLACK
    ws2.cell(row=row, column=1).border = BORDER
    for col in DAY_COLS:
        cell = ws2.cell(row=row, column=col)
        cell.value = hours_formula(name, i, col)
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

ws2.cell(row=CHECK_ROW, column=1, value="Έλεγχος Ανοίγματος (≥1)").font = SUBHEAD_FONT
ws2.cell(row=CHECK_ROW + 1, column=1, value="Έλεγχος Κλεισίματος (≥3)").font = SUBHEAD_FONT
for col in DAY_COLS:
    col_letter = day_col_letter(col)
    open_check = ws2.cell(row=CHECK_ROW, column=col, value=f'=COUNTIF({col_letter}{EFFECTIVE_SHIFT_ROW}:{col_letter}{EFFECTIVE_SHIFT_ROW + 5},"Άνοιγμα")')
    close_check = ws2.cell(row=CHECK_ROW + 1, column=col, value=f'=COUNTIF({col_letter}{EFFECTIVE_SHIFT_ROW}:{col_letter}{EFFECTIVE_SHIFT_ROW + 5},"Κλείσιμο")')
    open_check.alignment = Alignment(horizontal="center")
    close_check.alignment = Alignment(horizontal="center")
    open_check.border = BORDER
    close_check.border = BORDER

shift_range = f"C{EFFECTIVE_SHIFT_ROW}:I{EFFECTIVE_SHIFT_ROW + 5}"
hours_range = f"C{HOURS_START_ROW}:I{HOURS_START_ROW + 5}"
open_check_range = f"C{CHECK_ROW}:I{CHECK_ROW}"
close_check_range = f"C{CHECK_ROW + 1}:I{CHECK_ROW + 1}"

ws2.conditional_formatting.add(
    shift_range,
    FormulaRule(formula=[f'C{EFFECTIVE_SHIFT_ROW}="Ρεπό"'], fill=REST_FILL),
)
ws2.conditional_formatting.add(
    shift_range,
    FormulaRule(formula=[f'C{EFFECTIVE_SHIFT_ROW}="Άνοιγμα"'], fill=LIGHT_FILL),
)
ws2.conditional_formatting.add(
    shift_range,
    FormulaRule(formula=[f'C{EFFECTIVE_SHIFT_ROW}="Κλείσιμο"'], fill=CLOSE_FILL),
)

hours_to_shift_offset = HOURS_START_ROW - EFFECTIVE_SHIFT_ROW
hours_shift_ref = f'INDIRECT(ADDRESS(ROW()-{hours_to_shift_offset},COLUMN()))'
for row in range(HOURS_START_ROW, HOURS_START_ROW + len(PEOPLE)):
    ws2.conditional_formatting.add(
        f"C{row}:I{row}",
        FormulaRule(formula=[f'{hours_shift_ref}="Ρεπό"'], fill=REST_FILL),
    )
    ws2.conditional_formatting.add(
        f"C{row}:I{row}",
        FormulaRule(formula=[f'{hours_shift_ref}="Άνοιγμα"'], fill=LIGHT_FILL),
    )
    ws2.conditional_formatting.add(
        f"C{row}:I{row}",
        FormulaRule(formula=[f'AND({hours_shift_ref}<>"Ρεπό",{hours_shift_ref}<>"Άνοιγμα")'], fill=CLOSE_FILL),
    )

ws2.conditional_formatting.add(
    open_check_range,
    FormulaRule(formula=[f"C{CHECK_ROW}<1"], fill=SHORTAGE_FILL, font=SHORTAGE_FONT),
)
ws2.conditional_formatting.add(
    open_check_range,
    FormulaRule(formula=[f"C{CHECK_ROW}>=1"], fill=OK_FILL),
)
ws2.conditional_formatting.add(
    close_check_range,
    FormulaRule(formula=[f"C{CHECK_ROW + 1}<3"], fill=SHORTAGE_FILL, font=SHORTAGE_FONT),
)
ws2.conditional_formatting.add(
    close_check_range,
    FormulaRule(formula=[f"C{CHECK_ROW + 1}>=3"], fill=OK_FILL),
)

output_path = Path(__file__).resolve().parent / "team_schedule_v2.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
